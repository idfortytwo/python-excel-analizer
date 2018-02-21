import re, sys, operator

from openpyxl import load_workbook
from pattern import Pattern
from json import encoder
from flask import Flask, request, jsonify
from werkzeug import secure_filename

encoder.FLOAT_REPR = lambda o: format(o, '.2f')

app = Flask(__name__)

patterns = {
    'YYYY MM DD HH':    Pattern(r'^\d{4}(/|-|\.| )(0?[1-9]|1[012])(/|-|\.| )(0?[1-9]|[12][0-9]|3[01]) (0?[0-9]|1[01]):(0?[0-9]|[1-5][0-9]) ([AaPp][Mm])$', 'DATE'),
    'YYYY MM DD HH24':  Pattern(r'^\d{4}(/|-|\.| )(0?[1-9]|1[012])(/|-|\.| )(0?[1-9]|[12][0-9]|3[01]) (0?[0-9]|1[0-9]|2[0-4]):(0?[0-9]|[1-5][0-9])$', 'DATE'),
    'DD MM YYYY':       Pattern(r'^(0?[1-9]|[12][0-9]|3[01])(/|-|\.| )(0?[1-9]|1[012])(/|-|\.| )\d{4}$', 'DATE'),
    'MM DD YYYY':       Pattern(r'^(0?[1-9]|1[012])(/|-|\.| )(0?[1-9]|[12][0-9]|3[01])(/|-|\.| )\d{4}$', 'DATE'),
    'YYYY MM DD':       Pattern(r'^\d{4}(/|-|\.| )(0?[1-9]|1[012])(/|-|\.| )(0?[1-9]|[12][0-9]|3[01])$', 'DATE'),
    'YYYY DD MM':       Pattern(r'^\d{4}(/|-|\.| )(0?[1-9]|[12][0-9]|3[01])(/|-|\.| )(0?[1-9]|1[012])$', 'DATE'),
    'DD MON YYYY':      Pattern(r'^(0?[1-9]|[12][0-9]|3[01])(/|-|\.| )[a-zA-Z]{3}(/|-|\.| )\d{4}$', 'DATE'),

    '.decimal':         Pattern(r'^\.\d+', 'NUMBER'),
    'comma decimal':    Pattern(r'\d{1,3}(,\d{3})+(\.\d+)$', 'NUMBER'),
    'decimal':          Pattern(r'\d+(\.\d+)$', 'NUMBER'),
    'number':           Pattern(r'^\d+$', 'NUMBER'),

    'text':             Pattern(r'\w*', 'VARCHAR')
}


def percentage(part, whole):
    return(round((100 * float(part) / float(whole)), 2))


@app.route('/', methods=['POST'])
def parse():

    file = open('temp.xlsx', 'wb')
    file.write(request.data)
    file.close()

    wb = load_workbook('temp.xlsx')
    name = wb.sheetnames[0]
    ws = wb[name]

    metadata = []
    for col in ws.iter_cols():

        frequency = {}
        for name in patterns:
            frequency.update({name:[0, '']})

        pc = 0.0
        for name, p in patterns.items():
            pattern = p.pattern
            type = p.type

            before_max = 0
            after_max = 0

            row_count = ws.max_row - 1

            for i, cell in enumerate(col):
                if i > 0:
                    value = str(cell.value)

                    if re.match(pattern, value):
                        frequency[name][0] += 1

                        # print('{} {} | {}'.format(type, name, value))

                        if type == 'NUMBER':
                            split = re.split(r'\.', value)
                            before = len(split[0])

                            if before > before_max:
                                before_max = before

                            if name != 'number':
                                after = len(split[1])
                                if after > after_max:
                                    after_max = after

                        elif type == 'VARCHAR':
                            length = len(value)
                            if length > before_max:
                                before_max = length

            if type == 'NUMBER' and name != 'number' and (before_max or after_max != 0):

                format = '9' * before_max + '.' + '9' * after_max
                frequency[name][1] = format

            elif type == 'NUMBER' and name == 'number':
                format = '9' * before_max
                frequency[name][1] = format

            elif type == 'VARCHAR':
                format = str(before_max)
                frequency[name][1] = format

            elif type == 'DATE':
                format = name
                frequency[name][1] = format

            pc = percentage(int(frequency[name][0]), row_count)
            if pc == 100.0:
                break;

        temp_values = []
        for name, value in frequency.items():
            count = value[0]
            format = value[1]
            if count > 0:
                temp_values.append([count, format])

        values = sorted(temp_values)[-2:]

        temp_frequency = frequency.copy()
        for name, frvalue in frequency.items():
            if not frvalue in values:
                del temp_frequency[name]
        frequency = temp_frequency.copy()

        # print('temp_counts: {}'.format(temp_counts))
        # print('temp_values: {}'.format(temp_values))
        # print('values: {}'.format(values))
        # print('counts: {}'.format(counts))
        # print('frequency: {}'.format(frequency))

        # print(list(frequency.items())[0])
        # print(list(frequency.items())[1])

        if len(frequency) == 1:

            key = list(frequency.keys())[0]
            value = list(frequency.values())[0]

            type = patterns[key].type
            count = value[0]

            if type == 'VARCHAR':
                length = int(value[1])
                format = None
            else:
                length = None
                format = value[1]

            column = {
                'column_name': col[0].value,
                'data_type': type,
                'data_length': length,
                'data_format': format,
                'percentage': pc
            }
            metadata.append(column)

        elif len(frequency) == 2:
            key = list(frequency.keys())[0]
            value = list(frequency.values())[0]

            type = patterns[key].type
            count = value[0]
            pc = percentage(int(count), row_count)

            if type == 'VARCHAR':
                length = int(value[1])
                format = None
            else:
                length = None
                format = value[1]

            alternative_column = {
                'data_type': type,
                'data_length': length,
                'data_format': format,
                'percentage': pc
            }

            key = list(frequency.keys())[1]
            value = list(frequency.values())[1]

            type = patterns[key].type
            count = value[0]
            pc = percentage(int(count), row_count)

            if type == 'VARCHAR':
                length = int(value[1])
                format = None
            else:
                length = None
                format = value[1]

            column = {
                'column_name': col[0].value,
                'data_type': type,
                'data_length': length,
                'data_format': format,
                'percentage': pc,
                'alternative_column': alternative_column
            }
            metadata.append(column)

    data = []
    for i, r in enumerate(ws.iter_rows()):
        if i > 0 and i < 100:
            row = {}
            for cell in r:
                row.update({ws.cell(row=1, column=cell.col_idx).value: cell.value})
            i += 1

            data.append(row)


    return jsonify(metadata=metadata, data=data)


if __name__ == '__main__':
    app.run(debug=True, host= '0.0.0.0')
